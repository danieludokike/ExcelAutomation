from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


class ExcelAutomator:
    def __init__(self, filename=None):
        """Gets the filename and sets it"""
        self._filename = filename

    def read_file(self):
        try:
            wb = load_workbook(self._filename)
        except FileNotFoundError as e:
            raise FileNotFoundError(f"FILE_ERROR [{e}]")

        #       return worksheet object
        ws = wb.active
        # ws.title = "New Title"
        return ws

    def get_max_row_and_col(self):
        worksheet = self.read_file()
        return worksheet.max_row, worksheet.max_column

    def print_file_values(self):
        """PRINTS THE FILE VALUE"""
        row_num, col_num = self.get_max_row_and_col()
        ws = self.read_file()

        for row in range(1, row_num + 1):
            for col in range(1, col_num + 1):
                char = get_column_letter(col)
                print(ws[char + str(row)].value)

        return None

    def save_excels_sheet(self, name_to_save):
        #         Current worksheet
        ws = self.read_file()
        ws.save(name_to_save)
        return None


if __name__ == "__main__":
    ex_automator = ExcelAutomator("ggdaniel.xlsx")
    ex_automator.print_file_values()
    print(ex_automator.get_max_row_and_col())

# wb = load_workbook("daniel.xlsx")
# ws = wb.active

# Looping through cells
# for row in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)


# Merging cells in openpyxl
# ws.merged_cells("A1:D1")
# ws.save("daniel.xlsx")

# wb = load_workbook("IT.xlsx")
# ws = wb.active
# print(ws["D9"].value)
# ws['D9'].value = "Test"
# wb.save('IT.xlsx')

# wb.create_sheet("Test")
# #
# # print(wb.sheetnames)
# #
# # wb.save("IT.xlsx")
#
# # Create a new workbook
# wb = Workbook()
# # Get an active work sheet
# ws = wb.active
# # Change the Title of the worksheet
# ws.title = "Data"
# print(ws.title)
#
# # Appending Data to the worksheet
# ws.append(("Daniel", "is", "A", "Good"))
# ws.append(("Daniel", "is", "A", "Good"))
# print(ws[2][-1].value)
# # wb.save("daniel.xlsx")
